import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# this test is to show how to extract data from excel sheet
book = openpyxl.load_workbook(os.path.join(BASE_DIR, "TestData", "Input data", "PythonDemo.xlsx"))
sheet = book.active
Dict ={}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Swati"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet["A5"].value)

