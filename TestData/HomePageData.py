import openpyxl
import os

class HomePageData:

    test_homepage_data = [{"firstname":"Swati", "lastname":"Doshi", "gender":"Female"}, {"firstname":"Ankit", "lastname":"Doshi", "gender":"Male"}]

    @staticmethod
    def get_TestData(test_case_name):
        Dict = {}
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        book = openpyxl.load_workbook(os.path.join(BASE_DIR, "TestData", "Input data", "PythonDemo.xlsx"))
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # to get data of a particular test case

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print([Dict])
        return [Dict]